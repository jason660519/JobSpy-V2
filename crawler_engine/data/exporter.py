"""數據導出器

提供多種格式的數據導出功能，包括JSON、CSV、Excel、PDF等格式。
"""

import json
import csv
import asyncio
from typing import Dict, List, Any, Optional, Union, IO
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from abc import ABC, abstractmethod
import structlog
from enum import Enum

logger = structlog.get_logger(__name__)


class ExportFormat(Enum):
    """導出格式"""
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    XML = "xml"
    PDF = "pdf"
    HTML = "html"
    MARKDOWN = "markdown"


class CompressionType(Enum):
    """壓縮類型"""
    NONE = "none"
    GZIP = "gzip"
    ZIP = "zip"
    BZIP2 = "bzip2"


@dataclass
class ExportConfig:
    """導出配置"""
    format: ExportFormat = ExportFormat.JSON
    output_path: Optional[str] = None
    compression: CompressionType = CompressionType.NONE
    encoding: str = "utf-8"
    include_metadata: bool = True
    pretty_print: bool = True
    batch_size: int = 1000
    max_file_size: int = 100 * 1024 * 1024  # 100MB
    
    # CSV特定配置
    csv_delimiter: str = ","
    csv_quotechar: str = '"'
    csv_include_header: bool = True
    
    # Excel特定配置
    excel_sheet_name: str = "JobData"
    excel_include_charts: bool = False
    
    # PDF特定配置
    pdf_page_size: str = "A4"
    pdf_orientation: str = "portrait"
    
    # HTML特定配置
    html_template: Optional[str] = None
    html_include_css: bool = True


@dataclass
class ExportStats:
    """導出統計"""
    total_records: int = 0
    exported_records: int = 0
    failed_records: int = 0
    files_created: int = 0
    total_size: int = 0
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    errors: List[str] = field(default_factory=list)
    
    @property
    def success_rate(self) -> float:
        """成功率"""
        if self.total_records == 0:
            return 0.0
        return self.exported_records / self.total_records
    
    @property
    def duration(self) -> Optional[float]:
        """導出耗時（秒）"""
        if self.start_time and self.end_time:
            return (self.end_time - self.start_time).total_seconds()
        return None


@dataclass
class ExportMetadata:
    """導出元數據"""
    export_time: datetime
    format: str
    total_records: int
    source: str
    version: str = "1.0"
    filters: Optional[Dict[str, Any]] = None
    schema_version: str = "1.0"


class DataExporter(ABC):
    """數據導出器基類"""
    
    def __init__(self, config: ExportConfig):
        self.config = config
        self.stats = ExportStats()
        self.logger = logger.bind(exporter=self.__class__.__name__)
    
    @abstractmethod
    async def export(self, data: List[Dict[str, Any]], output_path: str) -> bool:
        """導出數據
        
        Args:
            data: 要導出的數據
            output_path: 輸出路徑
            
        Returns:
            bool: 是否成功
        """
        pass
    
    @abstractmethod
    def get_file_extension(self) -> str:
        """獲取文件擴展名"""
        pass
    
    def _prepare_metadata(self, data: List[Dict[str, Any]], source: str = "JobSpy") -> ExportMetadata:
        """準備元數據"""
        return ExportMetadata(
            export_time=datetime.utcnow(),
            format=self.config.format.value,
            total_records=len(data),
            source=source
        )
    
    def _should_split_file(self, current_size: int) -> bool:
        """檢查是否需要分割文件"""
        return current_size >= self.config.max_file_size
    
    def _get_split_filename(self, base_path: str, index: int) -> str:
        """獲取分割文件名"""
        path = Path(base_path)
        stem = path.stem
        suffix = path.suffix
        return str(path.parent / f"{stem}_part{index:03d}{suffix}")


class JSONExporter(DataExporter):
    """JSON導出器"""
    
    async def export(self, data: List[Dict[str, Any]], output_path: str) -> bool:
        """導出為JSON格式"""
        try:
            self.stats.start_time = datetime.utcnow()
            self.stats.total_records = len(data)
            
            # 準備導出數據
            export_data = {
                "data": data
            }
            
            if self.config.include_metadata:
                export_data["metadata"] = self._prepare_metadata(data).__dict__
            
            # 寫入文件
            with open(output_path, 'w', encoding=self.config.encoding) as f:
                if self.config.pretty_print:
                    json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
                else:
                    json.dump(export_data, f, ensure_ascii=False, default=str)
            
            self.stats.exported_records = len(data)
            self.stats.files_created = 1
            self.stats.total_size = Path(output_path).stat().st_size
            self.stats.end_time = datetime.utcnow()
            
            self.logger.info(
                "JSON導出完成",
                records=len(data),
                file=output_path,
                size=self.stats.total_size
            )
            
            return True
            
        except Exception as e:
            error_msg = f"JSON導出失敗: {str(e)}"
            self.stats.errors.append(error_msg)
            self.logger.error(error_msg, error=str(e))
            return False
    
    def get_file_extension(self) -> str:
        return ".json"


class CSVExporter(DataExporter):
    """CSV導出器"""
    
    async def export(self, data: List[Dict[str, Any]], output_path: str) -> bool:
        """導出為CSV格式"""
        try:
            self.stats.start_time = datetime.utcnow()
            self.stats.total_records = len(data)
            
            if not data:
                self.logger.warning("沒有數據可導出")
                return True
            
            # 獲取所有字段
            fieldnames = set()
            for record in data:
                fieldnames.update(record.keys())
            fieldnames = sorted(list(fieldnames))
            
            # 寫入CSV文件
            with open(output_path, 'w', newline='', encoding=self.config.encoding) as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=fieldnames,
                    delimiter=self.config.csv_delimiter,
                    quotechar=self.config.csv_quotechar,
                    quoting=csv.QUOTE_MINIMAL
                )
                
                if self.config.csv_include_header:
                    writer.writeheader()
                
                for record in data:
                    # 處理嵌套對象
                    flattened_record = self._flatten_record(record)
                    writer.writerow(flattened_record)
            
            self.stats.exported_records = len(data)
            self.stats.files_created = 1
            self.stats.total_size = Path(output_path).stat().st_size
            self.stats.end_time = datetime.utcnow()
            
            self.logger.info(
                "CSV導出完成",
                records=len(data),
                file=output_path,
                size=self.stats.total_size
            )
            
            return True
            
        except Exception as e:
            error_msg = f"CSV導出失敗: {str(e)}"
            self.stats.errors.append(error_msg)
            self.logger.error(error_msg, error=str(e))
            return False
    
    def _flatten_record(self, record: Dict[str, Any], prefix: str = "") -> Dict[str, str]:
        """扁平化記錄"""
        flattened = {}
        
        for key, value in record.items():
            new_key = f"{prefix}{key}" if prefix else key
            
            if isinstance(value, dict):
                flattened.update(self._flatten_record(value, f"{new_key}."))
            elif isinstance(value, list):
                # 將列表轉換為字符串
                flattened[new_key] = "; ".join(str(item) for item in value)
            else:
                flattened[new_key] = str(value) if value is not None else ""
        
        return flattened
    
    def get_file_extension(self) -> str:
        return ".csv"


class ExcelExporter(DataExporter):
    """Excel導出器"""
    
    async def export(self, data: List[Dict[str, Any]], output_path: str) -> bool:
        """導出為Excel格式"""
        try:
            # 嘗試導入openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                error_msg = "需要安裝openpyxl庫才能導出Excel格式"
                self.stats.errors.append(error_msg)
                self.logger.error(error_msg)
                return False
            
            self.stats.start_time = datetime.utcnow()
            self.stats.total_records = len(data)
            
            if not data:
                self.logger.warning("沒有數據可導出")
                return True
            
            # 創建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.config.excel_sheet_name
            
            # 獲取所有字段
            fieldnames = set()
            for record in data:
                fieldnames.update(record.keys())
            fieldnames = sorted(list(fieldnames))
            
            # 寫入標題行
            for col, fieldname in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col, value=fieldname)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 寫入數據行
            for row, record in enumerate(data, 2):
                for col, fieldname in enumerate(fieldnames, 1):
                    value = record.get(fieldname, "")
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    ws.cell(row=row, column=col, value=value)
            
            # 自動調整列寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 添加元數據工作表
            if self.config.include_metadata:
                metadata_ws = wb.create_sheet("Metadata")
                metadata = self._prepare_metadata(data)
                
                metadata_ws.cell(row=1, column=1, value="屬性")
                metadata_ws.cell(row=1, column=2, value="值")
                
                row = 2
                for key, value in metadata.__dict__.items():
                    metadata_ws.cell(row=row, column=1, value=key)
                    metadata_ws.cell(row=row, column=2, value=str(value))
                    row += 1
            
            # 保存文件
            wb.save(output_path)
            
            self.stats.exported_records = len(data)
            self.stats.files_created = 1
            self.stats.total_size = Path(output_path).stat().st_size
            self.stats.end_time = datetime.utcnow()
            
            self.logger.info(
                "Excel導出完成",
                records=len(data),
                file=output_path,
                size=self.stats.total_size
            )
            
            return True
            
        except Exception as e:
            error_msg = f"Excel導出失敗: {str(e)}"
            self.stats.errors.append(error_msg)
            self.logger.error(error_msg, error=str(e))
            return False
    
    def get_file_extension(self) -> str:
        return ".xlsx"


class HTMLExporter(DataExporter):
    """HTML導出器"""
    
    async def export(self, data: List[Dict[str, Any]], output_path: str) -> bool:
        """導出為HTML格式"""
        try:
            self.stats.start_time = datetime.utcnow()
            self.stats.total_records = len(data)
            
            if not data:
                self.logger.warning("沒有數據可導出")
                return True
            
            # 獲取所有字段
            fieldnames = set()
            for record in data:
                fieldnames.update(record.keys())
            fieldnames = sorted(list(fieldnames))
            
            # 生成HTML內容
            html_content = self._generate_html(data, fieldnames)
            
            # 寫入文件
            with open(output_path, 'w', encoding=self.config.encoding) as f:
                f.write(html_content)
            
            self.stats.exported_records = len(data)
            self.stats.files_created = 1
            self.stats.total_size = Path(output_path).stat().st_size
            self.stats.end_time = datetime.utcnow()
            
            self.logger.info(
                "HTML導出完成",
                records=len(data),
                file=output_path,
                size=self.stats.total_size
            )
            
            return True
            
        except Exception as e:
            error_msg = f"HTML導出失敗: {str(e)}"
            self.stats.errors.append(error_msg)
            self.logger.error(error_msg, error=str(e))
            return False
    
    def _generate_html(self, data: List[Dict[str, Any]], fieldnames: List[str]) -> str:
        """生成HTML內容"""
        css_styles = """
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #f2f2f2; font-weight: bold; }
            tr:nth-child(even) { background-color: #f9f9f9; }
            .metadata { margin-bottom: 20px; padding: 10px; background-color: #e9e9e9; }
        </style>
        """ if self.config.html_include_css else ""
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="{self.config.encoding}">
            <title>JobSpy 數據導出</title>
            {css_styles}
        </head>
        <body>
            <h1>JobSpy 數據導出</h1>
        """
        
        # 添加元數據
        if self.config.include_metadata:
            metadata = self._prepare_metadata(data)
            html += f"""
            <div class="metadata">
                <h2>導出信息</h2>
                <p><strong>導出時間:</strong> {metadata.export_time}</p>
                <p><strong>記錄總數:</strong> {metadata.total_records}</p>
                <p><strong>格式:</strong> {metadata.format}</p>
                <p><strong>來源:</strong> {metadata.source}</p>
            </div>
            """
        
        # 添加數據表格
        html += """
            <table>
                <thead>
                    <tr>
        """
        
        for fieldname in fieldnames:
            html += f"<th>{fieldname}</th>"
        
        html += """
                    </tr>
                </thead>
                <tbody>
        """
        
        for record in data:
            html += "<tr>"
            for fieldname in fieldnames:
                value = record.get(fieldname, "")
                if isinstance(value, (list, dict)):
                    value = str(value)
                html += f"<td>{value}</td>"
            html += "</tr>"
        
        html += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        return html
    
    def get_file_extension(self) -> str:
        return ".html"


class DataExportManager:
    """數據導出管理器
    
    統一管理各種格式的數據導出功能。
    """
    
    def __init__(self):
        self.logger = logger.bind(component="DataExportManager")
        self._exporters = {
            ExportFormat.JSON: JSONExporter,
            ExportFormat.CSV: CSVExporter,
            ExportFormat.EXCEL: ExcelExporter,
            ExportFormat.HTML: HTMLExporter,
        }
    
    def register_exporter(self, format: ExportFormat, exporter_class: type) -> None:
        """註冊導出器
        
        Args:
            format: 導出格式
            exporter_class: 導出器類
        """
        self._exporters[format] = exporter_class
        self.logger.info("註冊導出器", format=format.value, exporter=exporter_class.__name__)
    
    def get_supported_formats(self) -> List[ExportFormat]:
        """獲取支持的格式"""
        return list(self._exporters.keys())
    
    async def export_data(self, 
                         data: List[Dict[str, Any]], 
                         config: ExportConfig,
                         output_path: Optional[str] = None) -> bool:
        """導出數據
        
        Args:
            data: 要導出的數據
            config: 導出配置
            output_path: 輸出路徑
            
        Returns:
            bool: 是否成功
        """
        try:
            # 檢查格式支持
            if config.format not in self._exporters:
                self.logger.error("不支持的導出格式", format=config.format.value)
                return False
            
            # 創建導出器
            exporter_class = self._exporters[config.format]
            exporter = exporter_class(config)
            
            # 確定輸出路徑
            if output_path is None:
                if config.output_path:
                    output_path = config.output_path
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"jobspy_export_{timestamp}{exporter.get_file_extension()}"
                    output_path = str(Path.cwd() / filename)
            
            # 確保輸出目錄存在
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            # 執行導出
            success = await exporter.export(data, output_path)
            
            if success:
                self.logger.info(
                    "數據導出成功",
                    format=config.format.value,
                    records=len(data),
                    file=output_path
                )
            else:
                self.logger.error(
                    "數據導出失敗",
                    format=config.format.value,
                    errors=exporter.stats.errors
                )
            
            return success
            
        except Exception as e:
            self.logger.error("導出過程中發生錯誤", error=str(e))
            return False
    
    async def export_multiple_formats(self, 
                                     data: List[Dict[str, Any]], 
                                     formats: List[ExportFormat],
                                     base_config: ExportConfig,
                                     output_dir: Optional[str] = None) -> Dict[ExportFormat, bool]:
        """導出多種格式
        
        Args:
            data: 要導出的數據
            formats: 導出格式列表
            base_config: 基礎配置
            output_dir: 輸出目錄
            
        Returns:
            Dict[ExportFormat, bool]: 各格式的導出結果
        """
        results = {}
        
        if output_dir:
            Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        for format in formats:
            try:
                # 創建格式特定的配置
                config = ExportConfig(
                    format=format,
                    output_path=base_config.output_path,
                    compression=base_config.compression,
                    encoding=base_config.encoding,
                    include_metadata=base_config.include_metadata,
                    pretty_print=base_config.pretty_print,
                    batch_size=base_config.batch_size,
                    max_file_size=base_config.max_file_size
                )
                
                # 生成輸出路徑
                if output_dir:
                    exporter_class = self._exporters[format]
                    exporter = exporter_class(config)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"jobspy_export_{timestamp}{exporter.get_file_extension()}"
                    output_path = str(Path(output_dir) / filename)
                else:
                    output_path = None
                
                # 執行導出
                success = await self.export_data(data, config, output_path)
                results[format] = success
                
            except Exception as e:
                self.logger.error(
                    "格式導出失敗",
                    format=format.value,
                    error=str(e)
                )
                results[format] = False
        
        return results
    
    async def batch_export(self, 
                          data: List[Dict[str, Any]], 
                          config: ExportConfig,
                          output_dir: str) -> List[str]:
        """批量導出（分割大文件）
        
        Args:
            data: 要導出的數據
            config: 導出配置
            output_dir: 輸出目錄
            
        Returns:
            List[str]: 生成的文件路徑列表
        """
        output_files = []
        
        try:
            Path(output_dir).mkdir(parents=True, exist_ok=True)
            
            # 計算批次數量
            batch_size = config.batch_size
            total_batches = (len(data) + batch_size - 1) // batch_size
            
            for i in range(total_batches):
                start_idx = i * batch_size
                end_idx = min((i + 1) * batch_size, len(data))
                batch_data = data[start_idx:end_idx]
                
                # 生成批次文件名
                exporter_class = self._exporters[config.format]
                exporter = exporter_class(config)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"jobspy_batch_{i+1:03d}_{timestamp}{exporter.get_file_extension()}"
                output_path = str(Path(output_dir) / filename)
                
                # 導出批次數據
                success = await self.export_data(batch_data, config, output_path)
                
                if success:
                    output_files.append(output_path)
                    self.logger.info(
                        "批次導出完成",
                        batch=i+1,
                        total_batches=total_batches,
                        records=len(batch_data),
                        file=output_path
                    )
                else:
                    self.logger.error(
                        "批次導出失敗",
                        batch=i+1,
                        total_batches=total_batches
                    )
            
            return output_files
            
        except Exception as e:
            self.logger.error("批量導出失敗", error=str(e))
            return output_files